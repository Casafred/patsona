"""Excel 样本导入 - 从 Excel 文件加载已分类的样本专利"""

from dataclasses import dataclass
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook


@dataclass
class SamplePatent:
    """样本专利数据"""

    # 专利号/申请号
    patent_id: str
    # 发明名称
    title: str
    # 摘要
    abstract: str
    # 所属技术分支ID
    branch_id: str


# 默认列映射配置
# 键为 SamplePatent 字段名，值为可能的列标题（按优先级排列）
DEFAULT_COLUMN_MAPPING: dict[str, list[str]] = {
    "patent_id": ["专利号", "申请号", "专利ID", "patent_id", "id"],
    "title": ["标题", "发明名称", "专利名称", "title", "name"],
    "abstract": ["摘要", "专利摘要", "abstract", "summary"],
    "branch_id": ["分类", "分类标签", "分支ID", "branch_id", "category", "label"],
}


class ExcelSampleLoader:
    """Excel 样本加载器

    从 Excel 文件加载已分类的样本专利数据，
    用于 Layer3 细分类时的样本对比参考。

    支持自动列映射：根据列标题自动匹配到对应字段。
    """

    def __init__(
        self,
        column_mapping: Optional[dict[str, list[str]]] = None,
    ) -> None:
        """初始化加载器

        Args:
            column_mapping: 自定义列映射，None则使用默认映射
        """
        self.column_mapping = column_mapping or DEFAULT_COLUMN_MAPPING

    def load(
        self,
        path: Path,
        sheet_name: Optional[str] = None,
    ) -> list[SamplePatent]:
        """从 Excel 文件加载样本专利

        Args:
            path: Excel 文件路径
            sheet_name: 工作表名称，None则使用第一个工作表

        Returns:
            样本专利列表

        Raises:
            FileNotFoundError: 文件不存在
            ValueError: 必要列未找到
        """
        path = Path(path)
        if not path.exists():
            raise FileNotFoundError(f"文件不存在: {path}")

        # 打开工作簿
        wb = load_workbook(str(path), read_only=True, data_only=True)

        # 选择工作表
        if sheet_name:
            ws = wb[sheet_name]
        else:
            ws = wb.active

        if ws is None:
            wb.close()
            return []

        # 读取表头行
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            wb.close()
            return []

        headers = [str(h).strip() if h else "" for h in rows[0]]

        # 建立列映射：字段名 -> 列索引
        col_index = self._build_column_index(headers)

        # 检查必要字段
        required_fields = ["patent_id", "branch_id"]
        for field_name in required_fields:
            if field_name not in col_index:
                wb.close()
                raise ValueError(
                    f"未找到必要列: {field_name}，"
                    f"请确保表头包含 {self.column_mapping[field_name]} 之一"
                )

        # 读取数据行
        samples: list[SamplePatent] = []
        for row in rows[1:]:
            if not row or all(cell is None for cell in row):
                continue

            sample = self._parse_row(row, col_index)
            if sample and sample.patent_id and sample.branch_id:
                samples.append(sample)

        wb.close()
        return samples

    def _build_column_index(self, headers: list[str]) -> dict[str, int]:
        """根据表头建立字段到列索引的映射

        Args:
            headers: 表头行数据

        Returns:
            字段名 -> 列索引 的映射字典
        """
        col_index: dict[str, int] = {}

        for field_name, possible_names in self.column_mapping.items():
            for col_idx, header in enumerate(headers):
                header_lower = header.lower().strip()
                for name in possible_names:
                    if header_lower == name.lower():
                        col_index[field_name] = col_idx
                        break
                if field_name in col_index:
                    break

        return col_index

    def _parse_row(
        self, row: tuple, col_index: dict[str, int]
    ) -> Optional[SamplePatent]:
        """解析单行数据为 SamplePatent

        Args:
            row: 行数据元组
            col_index: 字段到列索引的映射

        Returns:
            SamplePatent 对象，解析失败返回 None
        """
        def get_value(field_name: str) -> str:
            """安全获取字段值"""
            idx = col_index.get(field_name)
            if idx is not None and idx < len(row):
                val = row[idx]
                return str(val).strip() if val is not None else ""
            return ""

        patent_id = get_value("patent_id")
        title = get_value("title")
        abstract = get_value("abstract")
        branch_id = get_value("branch_id")

        if not patent_id or not branch_id:
            return None

        return SamplePatent(
            patent_id=patent_id,
            title=title,
            abstract=abstract,
            branch_id=branch_id,
        )
