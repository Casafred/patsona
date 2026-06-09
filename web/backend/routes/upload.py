"""文件上传API - 处理PDF、Word、TXT、Excel文件上传"""

import json
import tempfile
from pathlib import Path
from typing import Optional

from fastapi import APIRouter, File, Form, UploadFile, HTTPException
from pydantic import BaseModel

# 导入patsona模块
import sys
sys.path.insert(0, str(Path(__file__).parent.parent.parent.parent))

from openpyxl import load_workbook

router = APIRouter(prefix="/api/upload", tags=["upload"])


class ExcelInfoResponse(BaseModel):
    """Excel信息响应"""
    success: bool
    columns: list[str] = []
    row_count: int = 0
    sample_data: list[dict] = []
    error: Optional[str] = None


class ExcelProcessResponse(BaseModel):
    """Excel处理响应"""
    success: bool
    texts: list[str] = []
    row_count: int = 0
    error: Optional[str] = None


@router.post("/excel-info", response_model=ExcelInfoResponse)
async def get_excel_info(file: UploadFile = File(...)):
    """获取Excel文件的列信息和样本数据"""
    try:
        # 保存临时文件
        suffix = Path(file.filename).suffix.lower()
        if suffix not in ['.xlsx', '.xls']:
            return ExcelInfoResponse(success=False, error="不是Excel文件")

        with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
            content = await file.read()
            tmp.write(content)
            tmp_path = Path(tmp.name)

        # 读取Excel
        wb = load_workbook(tmp_path, read_only=True)
        ws = wb.active

        # 获取列名（第一行）
        columns = []
        for cell in ws[1]:
            if cell.value:
                columns.append(str(cell.value))

        # 获取行数和数据样本（前5行）
        row_count = ws.max_row - 1  # 减去标题行
        sample_data = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=6, values_only=True), start=2):
            if row_idx > 6:
                break
            row_dict = {}
            for col_idx, col_name in enumerate(columns):
                if col_idx < len(row):
                    row_dict[col_name] = str(row[col_idx] or '')
            sample_data.append(row_dict)

        wb.close()
        tmp_path.unlink(missing_ok=True)

        return ExcelInfoResponse(
            success=True,
            columns=columns,
            row_count=row_count,
            sample_data=sample_data
        )

    except Exception as e:
        return ExcelInfoResponse(success=False, error=str(e))


@router.post("/excel-process", response_model=ExcelProcessResponse)
async def process_excel(
    file: UploadFile = File(...),
    columns: str = Form(...),
    separator: str = Form(" ")
):
    """处理Excel文件，拼接指定列生成文本列表"""
    try:
        # 解析列配置
        selected_columns = json.loads(columns)
        sep = separator if separator != "\n" else "\n"

        # 保存临时文件
        suffix = Path(file.filename).suffix.lower()
        if suffix not in ['.xlsx', '.xls']:
            return ExcelProcessResponse(success=False, error="不是Excel文件")

        with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
            content = await file.read()
            tmp.write(content)
            tmp_path = Path(tmp.name)

        # 读取Excel
        wb = load_workbook(tmp_path, read_only=True)
        ws = wb.active

        # 获取列名和索引映射
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        col_name_to_idx = {}
        for idx, name in enumerate(header_row):
            if name:
                col_name_to_idx[str(name)] = idx

        # 检查选择的列是否存在
        missing_cols = [c for c in selected_columns if c not in col_name_to_idx]
        if missing_cols:
            return ExcelProcessResponse(
                success=False,
                error=f"列不存在: {missing_cols}"
            )

        # 获取选择的列索引
        selected_indices = [col_name_to_idx[c] for c in selected_columns]

        # 拼接每行数据
        texts = []
        row_count = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            parts = []
            for idx in selected_indices:
                if idx < len(row) and row[idx]:
                    parts.append(str(row[idx]))
            if parts:
                texts.append(sep.join(parts))
                row_count += 1

        wb.close()
        tmp_path.unlink(missing_ok=True)

        return ExcelProcessResponse(
            success=True,
            texts=texts,
            row_count=row_count
        )

    except json.JSONDecodeError:
        return ExcelProcessResponse(success=False, error="列配置格式错误")
    except Exception as e:
        return ExcelProcessResponse(success=False, error=str(e))


@router.post("/file")
async def upload_file(file: UploadFile = File(...)):
    """上传普通文件（PDF、Word、TXT）并解析"""
    try:
        suffix = Path(file.filename).suffix.lower()

        with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
            content = await file.read()
            tmp.write(content)
            tmp_path = Path(tmp.name)

        # 根据类型解析
        text = ""

        if suffix == '.txt':
            text = tmp_path.read_text(encoding='utf-8')

        elif suffix in ['.pdf']:
            # 使用patsona的解析器
            from patsona.preprocessor.parser import PatentParser
            parser = PatentParser()
            doc = parser.parse_pdf(tmp_path)
            text = doc.full_text

        elif suffix in ['.docx', '.doc']:
            from patsona.preprocessor.parser import PatentParser
            parser = PatentParser()
            doc = parser.parse_docx(tmp_path)
            text = doc.full_text

        tmp_path.unlink(missing_ok=True)

        return {
            "success": True,
            "text": text,
            "filename": file.filename
        }

    except Exception as e:
        return {"success": False, "error": str(e)}